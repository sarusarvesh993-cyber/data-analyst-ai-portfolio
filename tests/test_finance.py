"""Tests for Project 06 financial planning and scenario outputs."""
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook

from portfolio_app.finance import (
    apply_corporate_scenario,
    corporate_pnl,
    filter_planning_mart,
    load_outputs,
)

ROOT = Path(__file__).parents[1]
PROJECT = ROOT / "06-financial-planning"
OUTPUTS = PROJECT / "outputs"


def test_committed_finance_outputs_reconcile_to_reviewed_snapshot():
    outputs = load_outputs(OUTPUTS)
    assert set(outputs) == {
        "kpis",
        "departments",
        "funds",
        "expenses",
        "mart",
        "drivers",
        "corporate_plan",
        "corporate_monthly",
        "quality",
        "metadata",
    }
    kpis = outputs["kpis"].iloc[0]
    assert int(kpis["budget_fiscal_year"]) == 2026
    assert int(kpis["through_quarter"]) == 3
    assert int(kpis["source_rows"]) == 57_267
    assert kpis["annual_budget"] == pytest.approx(8_101_990_455.95)
    assert kpis["expenditures_to_date"] == pytest.approx(6_021_243_202.00)
    assert kpis["utilization_pct"] == pytest.approx(74.31807325295)
    assert len(outputs["departments"]) == 40
    assert outputs["departments"]["dept_rollup_name"].is_unique
    assert outputs["quality"].loc[
        outputs["quality"]["check_name"].eq("duplicate_key_rows"), "issue_count"
    ].iloc[0] == 0


def test_planning_filters_preserve_source_and_selected_grain():
    mart = load_outputs(OUTPUTS)["mart"]
    original_rows = len(mart)
    selected = filter_planning_mart(
        mart,
        departments=["Austin Energy"],
        funds=["Austin Energy Fund"],
    )
    assert len(mart) == original_rows
    assert not selected.empty
    assert selected["dept_rollup_name"].eq("Austin Energy").all()
    assert selected["fund_name"].eq("Austin Energy Fund").all()
    assert selected["program_name"].nunique() > 1


def test_corporate_scenarios_change_only_future_periods_and_are_directional():
    plan = load_outputs(OUTPUTS)["corporate_plan"]
    base = apply_corporate_scenario(plan)
    downside = apply_corporate_scenario(
        plan,
        revenue_adjustment_pct=-5,
        cost_inflation_pct=4,
        hiring_delay_savings_pct=0,
    )
    hiring_action = apply_corporate_scenario(
        plan,
        revenue_adjustment_pct=-5,
        cost_inflation_pct=4,
        hiring_delay_savings_pct=10,
    )
    actual = plan["period_status"].eq("Actual")
    assert downside.loc[actual, "scenario_amount"].equals(
        base.loc[actual, "scenario_amount"]
    )
    assert corporate_pnl(downside, "scenario_amount")["ebitda"] < corporate_pnl(
        base, "scenario_amount"
    )["ebitda"]
    assert corporate_pnl(hiring_action, "scenario_amount")["ebitda"] > corporate_pnl(
        downside, "scenario_amount"
    )["ebitda"]
    with pytest.raises(ValueError):
        apply_corporate_scenario(plan, revenue_adjustment_pct=50)


def test_excel_model_contains_editable_assumptions_and_formulas():
    workbook_path = PROJECT / "assets" / "project_06_fpa_model.xlsx"
    workbook = load_workbook(workbook_path, data_only=False, read_only=False)
    assert {
        "Read_Me",
        "Austin_Executive",
        "Departments",
        "Funds",
        "Variance_Drivers",
        "Data_Quality",
        "Scenario_Assumptions",
        "Corporate_Plan",
        "Corporate_Dashboard",
    }.issubset(workbook.sheetnames)
    assert workbook["Scenario_Assumptions"]["B4"].value == 0
    assert str(workbook["Corporate_Plan"]["I5"].value).startswith("=IF(")
    assert str(workbook["Corporate_Dashboard"]["B4"].value).startswith("=SUMIFS(")
