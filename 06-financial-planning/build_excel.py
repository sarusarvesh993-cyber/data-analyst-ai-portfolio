"""Build the reviewed Project 06 Excel FP&A workbook."""
from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

PROJECT_DIR = Path(__file__).resolve().parent
OUTPUT_DIR = PROJECT_DIR / "outputs"
ASSET_DIR = PROJECT_DIR / "assets"
TARGET = ASSET_DIR / "project_06_fpa_model.xlsx"

NAVY = "173B63"
TEAL = "0F8A7B"
GOLD = "F4A340"
PALE = "EAF7F4"
LIGHT = "F4F8F7"
WHITE = "FFFFFF"
RED = "E26D5A"
SLATE = "63788B"


def _title(sheet, title: str, subtitle: str) -> None:
    sheet.sheet_view.showGridLines = False
    sheet.merge_cells("A1:J1")
    sheet["A1"] = title
    sheet["A1"].font = Font(size=18, bold=True, color=WHITE)
    sheet["A1"].fill = PatternFill("solid", fgColor=NAVY)
    sheet["A1"].alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 30
    sheet.merge_cells("A2:J2")
    sheet["A2"] = subtitle
    sheet["A2"].font = Font(size=10, color=SLATE)
    sheet["A2"].alignment = Alignment(wrap_text=True, vertical="center")
    sheet.row_dimensions[2].height = 32


def _write_frame(sheet, frame: pd.DataFrame, start_row: int, table_name: str) -> None:
    for col_index, column in enumerate(frame.columns, 1):
        cell = sheet.cell(start_row, col_index, str(column))
        cell.font = Font(bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=TEAL)
        cell.alignment = Alignment(wrap_text=True)
    for row_index, row in enumerate(frame.itertuples(index=False), start_row + 1):
        for col_index, value in enumerate(row, 1):
            if pd.isna(value):
                value = None
            sheet.cell(row_index, col_index, value)
    end_row = start_row + len(frame)
    end_col = len(frame.columns)
    reference = f"A{start_row}:{sheet.cell(end_row, end_col).coordinate}"
    table = Table(displayName=table_name, ref=reference)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(table)
    sheet.freeze_panes = f"A{start_row + 1}"
    sheet.auto_filter.ref = reference
    for column_cells in sheet.iter_cols(min_row=start_row, max_row=end_row, max_col=end_col):
        width = min(max(len(str(cell.value or "")) for cell in column_cells) + 2, 34)
        sheet.column_dimensions[column_cells[0].column_letter].width = max(width, 11)


def _format_financial_columns(sheet, header_row: int, max_row: int) -> None:
    currency_keywords = {
        "budget",
        "expenditures",
        "expected_spend_to_date",
        "remaining_budget",
        "pace_variance",
        "linear_run_rate_proxy",
        "projected_variance_proxy",
        "absolute_pace_variance",
        "budget_amount",
        "actual_amount",
        "base_forecast_amount",
    }
    percent_keywords = {"elapsed_pct", "utilization_pct", "pace_gap_pct_points"}
    headers = {sheet.cell(header_row, col).value: col for col in range(1, sheet.max_column + 1)}
    for header, column in headers.items():
        if header in currency_keywords:
            for row in range(header_row + 1, max_row + 1):
                sheet.cell(row, column).number_format = '$#,##0;[Red]-$#,##0'
        elif header in percent_keywords:
            for row in range(header_row + 1, max_row + 1):
                sheet.cell(row, column).number_format = '0.0'


def build_workbook() -> Path:
    """Create a formula-driven workbook with public and synthetic planning views."""
    ASSET_DIR.mkdir(parents=True, exist_ok=True)
    kpis = pd.read_csv(OUTPUT_DIR / "executive_kpis.csv").iloc[0]
    departments = pd.read_csv(OUTPUT_DIR / "department_summary.csv")
    funds = pd.read_csv(OUTPUT_DIR / "fund_summary.csv")
    drivers = pd.read_csv(OUTPUT_DIR / "variance_drivers.csv").head(250)
    plan = pd.read_csv(OUTPUT_DIR / "corporate_plan.csv")
    monthly = pd.read_csv(OUTPUT_DIR / "corporate_monthly.csv")
    quality = pd.read_csv(OUTPUT_DIR / "data_quality.csv")

    workbook = Workbook()
    readme = workbook.active
    readme.title = "Read_Me"
    _title(
        readme,
        "PROJECT 06 · FP&A MODEL",
        "Real City of Austin FY2026 operating-budget pacing plus a clearly labeled synthetic corporate scenario model.",
    )
    notes = [
        ("Purpose", "Demonstrate budget-versus-actual analysis, pacing, variance drivers, scenarios, and Excel delivery."),
        ("Public source", "City of Austin Program Budget Operating Budget Vs Expense Raw Data, dataset g5k8-8sud."),
        ("Snapshot", f"FY{int(kpis['budget_fiscal_year'])} through Q{int(kpis['through_quarter'])}; {int(kpis['source_rows']):,} source rows."),
        ("Pacing boundary", "Expected spend to date is annual budget multiplied by elapsed quarter share. It is a monitoring proxy, not an accounting forecast."),
        ("Known source caveat", "Personnel budgets and actuals can appear in different leave/pay objects. Zero-budget spend is retained for review, not automatically called an overrun."),
        ("Corporate model", "The Corporate_Plan and Scenario_Assumptions sheets are seeded synthetic demonstrations and are not City of Austin data."),
        ("How to use", "Change the three yellow assumption cells. Scenario formulas and dashboard metrics recalculate automatically in Excel."),
    ]
    for row_index, (label, value) in enumerate(notes, 4):
        readme.cell(row_index, 1, label).font = Font(bold=True, color=TEAL)
        readme.cell(row_index, 2, value).alignment = Alignment(wrap_text=True, vertical="top")
    readme.column_dimensions["A"].width = 24
    readme.column_dimensions["B"].width = 105

    executive = workbook.create_sheet("Austin_Executive")
    _title(
        executive,
        "CITY OF AUSTIN · BUDGET PACING",
        "Expense pacing through Q3. Positive pace variance means expenditures are above a straight-line 75% benchmark.",
    )
    cards = [
        ("Annual budget", float(kpis["annual_budget"]), '$0.00,,,"B"'),
        ("Expenditures to date", float(kpis["expenditures_to_date"]), '$0.00,,,"B"'),
        ("Utilization", float(kpis["utilization_pct"]) / 100, "0.0%"),
        ("Pace variance", float(kpis["pace_variance"]), '$0.0,,"M"'),
        ("Remaining budget", float(kpis["remaining_budget"]), '$0.00,,,"B"'),
        ("Departments above pace", int(kpis["above_pace_departments"]), "0"),
    ]
    for index, (label, value, number_format) in enumerate(cards):
        column = 1 + (index % 3) * 3
        row = 4 + (index // 3) * 3
        executive.merge_cells(start_row=row, start_column=column, end_row=row, end_column=column + 1)
        executive.cell(row, column, label).font = Font(bold=True, color=SLATE)
        executive.merge_cells(start_row=row + 1, start_column=column, end_row=row + 1, end_column=column + 1)
        cell = executive.cell(row + 1, column, value)
        cell.font = Font(size=18, bold=True, color=NAVY)
        cell.number_format = number_format
        cell.fill = PatternFill("solid", fgColor=LIGHT)

    dept_top = departments.head(12)[["dept_rollup_name", "budget", "expenditures"]]
    _write_frame(executive, dept_top, 11, "ExecutiveDepartmentChartData")
    _format_financial_columns(executive, 11, 11 + len(dept_top))
    chart = BarChart()
    chart.type = "bar"
    chart.style = 10
    chart.title = "Top departments: annual budget vs Q3 expenditures"
    chart.y_axis.title = "Department"
    chart.x_axis.title = "US dollars"
    data = Reference(executive, min_col=2, max_col=3, min_row=11, max_row=11 + len(dept_top))
    categories = Reference(executive, min_col=1, min_row=12, max_row=11 + len(dept_top))
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    chart.height = 8
    chart.width = 15
    executive.add_chart(chart, "E11")
    for column in range(1, 11):
        executive.column_dimensions[get_column_letter(column)].width = 17

    for sheet_name, frame, table_name in [
        ("Departments", departments, "DepartmentSummary"),
        ("Funds", funds, "FundSummary"),
        ("Variance_Drivers", drivers, "VarianceDrivers"),
        ("Data_Quality", quality, "DataQualityChecks"),
    ]:
        sheet = workbook.create_sheet(sheet_name)
        _title(sheet, sheet_name.replace("_", " "), "Reviewed output generated by build_finance.py.")
        _write_frame(sheet, frame, 4, table_name)
        _format_financial_columns(sheet, 4, 4 + len(frame))
        if "pace_variance" in frame.columns:
            column = frame.columns.get_loc("pace_variance") + 1
            sheet.conditional_formatting.add(
                f"{sheet.cell(5, column).coordinate}:{sheet.cell(4 + len(frame), column).coordinate}",
                ColorScaleRule(
                    start_type="min",
                    start_color="63BE7B",
                    mid_type="percentile",
                    mid_value=50,
                    mid_color="FFEB84",
                    end_type="max",
                    end_color="F8696B",
                ),
            )

    assumptions = workbook.create_sheet("Scenario_Assumptions")
    _title(
        assumptions,
        "CORPORATE SCENARIO ASSUMPTIONS",
        "Yellow cells are editable. Adjustments apply only to forecast months August–December 2026.",
    )
    assumption_rows = [
        ("Revenue adjustment", 0.00, "Applied to future revenue"),
        ("Cost inflation", 0.00, "Applied to future COGS and operating expense"),
        ("Hiring-delay savings", 0.00, "Offsets future Product & Engineering and G&A costs"),
    ]
    for row_index, (label, value, explanation) in enumerate(assumption_rows, 4):
        assumptions.cell(row_index, 1, label).font = Font(bold=True, color=NAVY)
        assumptions.cell(row_index, 2, value).number_format = "0.0%"
        assumptions.cell(row_index, 2).fill = PatternFill("solid", fgColor="FFF2CC")
        assumptions.cell(row_index, 2).border = Border(
            left=Side(style="thin", color=GOLD),
            right=Side(style="thin", color=GOLD),
            top=Side(style="thin", color=GOLD),
            bottom=Side(style="thin", color=GOLD),
        )
        assumptions.cell(row_index, 3, explanation)
    assumptions.column_dimensions["A"].width = 28
    assumptions.column_dimensions["B"].width = 18
    assumptions.column_dimensions["C"].width = 70

    corporate = workbook.create_sheet("Corporate_Plan")
    _title(
        corporate,
        "SYNTHETIC CORPORATE PLAN",
        "Seeded demonstration only. Scenario Amount and P&L Variance are formula-driven from Scenario_Assumptions.",
    )
    corporate_frame = plan[
        [
            "month",
            "period_status",
            "business_unit",
            "statement_group",
            "line_item",
            "budget_amount",
            "actual_amount",
            "base_forecast_amount",
        ]
    ].copy()
    corporate_frame.columns = [
        "Month",
        "Period Status",
        "Business Unit",
        "Statement Group",
        "Line Item",
        "Budget Amount",
        "Actual Amount",
        "Base Forecast Amount",
    ]
    _write_frame(corporate, corporate_frame, 4, "CorporatePlan")
    scenario_col = 9
    variance_col = 10
    corporate.cell(4, scenario_col, "Scenario Amount")
    corporate.cell(4, variance_col, "P&L Variance to Budget")
    for cell in [corporate.cell(4, scenario_col), corporate.cell(4, variance_col)]:
        cell.font = Font(bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=TEAL)
    for row in range(5, 5 + len(corporate_frame)):
        corporate.cell(
            row,
            scenario_col,
            f'=IF(B{row}="Actual",H{row},IF(D{row}="Revenue",H{row}*(1+Scenario_Assumptions!$B$4),H{row}*(1+Scenario_Assumptions!$B$5)*IF(OR(E{row}="Product & engineering",E{row}="General & administrative"),1-Scenario_Assumptions!$B$6,1)))',
        )
        corporate.cell(
            row,
            variance_col,
            f'=(I{row}-F{row})*IF(D{row}="Revenue",1,-1)',
        )
        for column in range(6, 11):
            corporate.cell(row, column).number_format = '$#,##0;[Red]-$#,##0'
    corporate.tables["CorporatePlan"].ref = f"A4:J{4 + len(corporate_frame)}"
    corporate.freeze_panes = "A5"
    corporate.column_dimensions["A"].width = 13
    for column in ["B", "C", "D", "E"]:
        corporate.column_dimensions[column].width = 24
    for column in ["F", "G", "H", "I", "J"]:
        corporate.column_dimensions[column].width = 20

    dashboard = workbook.create_sheet("Corporate_Dashboard")
    _title(
        dashboard,
        "SYNTHETIC CORPORATE FP&A DASHBOARD",
        "Formula-driven annual plan and scenario view. Change assumptions to test revenue, inflation, and hiring timing.",
    )
    metric_formulas = [
        ("Budget revenue", '=SUMIFS(Corporate_Plan!$F:$F,Corporate_Plan!$D:$D,"Revenue")'),
        ("Scenario revenue", '=SUMIFS(Corporate_Plan!$I:$I,Corporate_Plan!$D:$D,"Revenue")'),
        ("Budget costs", '=SUMIFS(Corporate_Plan!$F:$F,Corporate_Plan!$D:$D,"<>Revenue")'),
        ("Scenario costs", '=SUMIFS(Corporate_Plan!$I:$I,Corporate_Plan!$D:$D,"<>Revenue")'),
        ("Budget EBITDA", "=B4-B6"),
        ("Scenario EBITDA", "=B5-B7"),
        ("Scenario EBITDA margin", "=IFERROR(B9/B5,0)"),
    ]
    for row_index, (label, formula) in enumerate(metric_formulas, 4):
        dashboard.cell(row_index, 1, label).font = Font(bold=True, color=SLATE)
        dashboard.cell(row_index, 2, formula).font = Font(size=14, bold=True, color=NAVY)
        dashboard.cell(row_index, 2).fill = PatternFill("solid", fgColor=LIGHT)
        dashboard.cell(row_index, 2).number_format = "0.0%" if "margin" in label else '$0.0,,"M"'
    dashboard.column_dimensions["A"].width = 28
    dashboard.column_dimensions["B"].width = 22

    monthly_frame = monthly[
        ["month", "budget_ebitda", "base_forecast_ebitda"]
    ].copy()
    monthly_frame.columns = ["Month", "Budget EBITDA", "Base Forecast EBITDA"]
    _write_frame(dashboard, monthly_frame, 13, "CorporateMonthlyChartData")
    for row in range(14, 14 + len(monthly_frame)):
        dashboard.cell(row, 1).number_format = "mmm-yy"
        dashboard.cell(row, 2).number_format = '$0.0,,"M"'
        dashboard.cell(row, 3).number_format = '$0.0,,"M"'
    line_chart = LineChart()
    line_chart.title = "Monthly EBITDA: plan vs base forecast"
    line_chart.y_axis.title = "US dollars"
    line_chart.x_axis.title = "Month"
    data = Reference(dashboard, min_col=2, max_col=3, min_row=13, max_row=13 + len(monthly_frame))
    categories = Reference(dashboard, min_col=1, min_row=14, max_row=13 + len(monthly_frame))
    line_chart.add_data(data, titles_from_data=True)
    line_chart.set_categories(categories)
    line_chart.height = 8
    line_chart.width = 15
    dashboard.add_chart(line_chart, "E4")

    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.calculation.calcMode = "auto"
    workbook.save(TARGET)
    print(f"Wrote {TARGET} ({TARGET.stat().st_size / 1_000:.1f} KB)")
    return TARGET


if __name__ == "__main__":
    build_workbook()
